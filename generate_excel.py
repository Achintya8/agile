import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def read_csv(file_path):
    data = []
    with open(file_path, 'r', newline='', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            data.append(row)
    return data

def build_excel():
    wb = openpyxl.Workbook()
    
    # 1. Sheet 1: Summary Dashboard
    ws_summary = wb.active
    ws_summary.title = "Summary Dashboard"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Style definitions
    font_title = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    font_section = Font(name="Segoe UI", size=11, bold=True, color="1F497D")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    font_regular = Font(name="Segoe UI", size=10)
    
    fill_title = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fill_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    fill_zebra = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
    fill_highlight = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # soft green
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Title Banner
    ws_summary.merge_cells("A1:E1")
    title_cell = ws_summary["A1"]
    title_cell.value = "Sustain-Agile Calorie Tracker Comparative Telemetry Dashboard"
    title_cell.font = font_title
    title_cell.fill = fill_title
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 40
    
    # 100 Images Section
    ws_summary["A3"] = "1. Yesterday's Run (100-Image Sprint Backlog)"
    ws_summary["A3"].font = font_section
    
    headers_100 = ["Metric", "Control (Standard Scrum)", "Sustain-Agile (Proposed)", "Savings / Improvement"]
    for col_idx, h in enumerate(headers_100, 1):
        cell = ws_summary.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws_summary.row_dimensions[4].height = 25
    
    metrics_100 = [
        ["Sustainability Velocity ($V_s$)", 1.5272, 499.2793, "326.9x Improvement"],
        ["Total Computational Tokens", 68925, 2784, "96.0% Reduction"],
        ["Total Cloud Operational Cost", 0.010048, 0.000657, "93.5% Cost Saved"],
        ["Total Carbon Footprint", 72023.18, 142.83, "99.8% Carbon Saved"],
        ["Semantic Cache Hits", "0 / 100 (0.0%)", "75 / 100 (75.0%)", "+75 Caches Bypassed"]
    ]
    
    for row_offset, row_data in enumerate(metrics_100, 5):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_offset, column=col_idx, value=val)
            cell.font = font_regular
            cell.border = thin_border
            if row_offset % 2 == 1:
                cell.fill = fill_zebra
            
            # Format numbers
            if col_idx in [2, 3]:
                if isinstance(val, float):
                    if "Cost" in row_data[0]:
                        cell.number_format = "$#,##0.000000"
                    else:
                        cell.number_format = "#,##0.0000" if "Velocity" in row_data[0] else "#,##0.00"
                elif isinstance(val, int):
                    cell.number_format = "#,##0"
            if col_idx == 4:
                cell.font = font_bold
                cell.fill = fill_highlight
                cell.alignment = Alignment(horizontal="center")
                
    # 20 Images Section
    ws_summary["A11"] = "2. Today's Run (20-Image Sprint Backlog with API Rotation)"
    ws_summary["A11"].font = font_section
    
    headers_20 = ["Metric", "Control (Standard Scrum)", "Sustain-Agile (Proposed)", "Savings / Improvement"]
    for col_idx, h in enumerate(headers_20, 1):
        cell = ws_summary.cell(row=12, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws_summary.row_dimensions[12].height = 25
    
    metrics_20 = [
        ["Sustainability Velocity ($V_s$)", 1.5138, 142.9435, "94.4x Improvement"],
        ["Total Computational Tokens", 13907, 1644, "88.2% Reduction"],
        ["Total Cloud Operational Cost", 0.002014, 0.000294, "85.4% Cost Saved"],
        ["Total Carbon Footprint", 14366.26, 84.19, "99.4% Carbon Saved"],
        ["Semantic Cache Hits", "0 / 20 (0.0%)", "5 / 20 (25.0%)", "+5 Caches Bypassed"]
    ]
    
    for row_offset, row_data in enumerate(metrics_20, 13):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_offset, column=col_idx, value=val)
            cell.font = font_regular
            cell.border = thin_border
            if row_offset % 2 == 1:
                cell.fill = fill_zebra
            
            # Format numbers
            if col_idx in [2, 3]:
                if isinstance(val, float):
                    if "Cost" in row_data[0]:
                        cell.number_format = "$#,##0.000000"
                    else:
                        cell.number_format = "#,##0.0000" if "Velocity" in row_data[0] else "#,##0.00"
                elif isinstance(val, int):
                    cell.number_format = "#,##0"
            if col_idx == 4:
                cell.font = font_bold
                cell.fill = fill_highlight
                cell.alignment = Alignment(horizontal="center")

    # Auto-adjust column widths for Summary tab
    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_summary.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # 2. Sheet 2: 100 Images Run
    ws_100 = wb.create_sheet(title="100 Images Run")
    ws_100.views.sheetView[0].showGridLines = True
    data_100 = read_csv("sprint_telemetry_100_images.csv")
    
    # 3. Sheet 3: 20 Images Run
    ws_20 = wb.create_sheet(title="20 Images Run")
    ws_20.views.sheetView[0].showGridLines = True
    data_20 = read_csv("sprint_telemetry_20_images.csv")
    
    # Populate raw data sheets
    for ws, data in [(ws_100, data_100), (ws_20, data_20)]:
        # Headers
        for col_idx, val in enumerate(data[0], 1):
            cell = ws.cell(row=1, column=col_idx, value=val)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[1].height = 28
        
        # Rows
        for row_idx, row in enumerate(data[1:], 2):
            for col_idx, val in enumerate(row, 1):
                # Try parsing as float or int for numerical sorting/aggregation in Excel
                parsed_val = val
                try:
                    if '.' in val:
                        parsed_val = float(val)
                    else:
                        parsed_val = int(val)
                except ValueError:
                    pass
                
                cell = ws.cell(row=row_idx, column=col_idx, value=parsed_val)
                cell.font = font_regular
                cell.border = thin_border
                if row_idx % 2 == 1:
                    cell.fill = fill_zebra
                
                # Format specific columns if numerical
                if isinstance(parsed_val, (float, int)):
                    if col_idx in [5, 6, 7]: # Tokens
                        cell.number_format = "#,##0"
                    elif col_idx in [8, 9, 10]: # Cost / Egress
                        cell.number_format = "$#,##0.000000"
                    elif col_idx in [11, 13]: # Carbon / File Size
                        cell.number_format = "#,##0.00"
                        
        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            for cell in col:
                cell_str = str(cell.value or '')
                # Don't let explanation column explode the width
                if len(cell_str) > 40:
                    cell_str = cell_str[:40]
                if len(cell_str) > max_len:
                    max_len = len(cell_str)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
    # Save combined workbook to both workspace and current brain folder
    output_local_path = "c:\\Users\\achin\\Desktop\\agile\\combined_sprint_telemetry.xlsx"
    output_brain_path = "C:\\Users\\achin\\.gemini\\antigravity-ide\\brain\\82bd357f-5220-4459-893b-20a4c91fa18f\\combined_sprint_telemetry.xlsx"
    
    for path in [output_local_path, output_brain_path]:
        wb.save(path)
        print(f"Combined Excel saved to {path}")

if __name__ == "__main__":
    build_excel()
